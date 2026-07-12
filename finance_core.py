from __future__ import annotations

import os
import re
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import pandas as pd
import pdfplumber
from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


class StatementParseError(ValueError):
    """PDF ekstre işlem tablosuna dönüştürülemediğinde fırlatılır."""


TRANSACTION_COLUMNS = ["Sayfa", "Tarih", "Açıklama", "Tutar", "Ek_Bilgi", "Islem_Tipi"]

CATEGORY_RULES: list[tuple[str, tuple[str, ...]]] = [
    ("Faiz / Masraf", ("FAIZ", "MASRAF", "KOMISYON")),
    ("Ulaşım", ("BELBIM", "UBER", "MARMARAY", "AKARYAKIT", "SHELL", "OPET", "METRO", "TAKSI")),
    ("Market", ("GETIRPERAK", "MIGROS", "CARREFOURSA", "MARKET", "BIM", "A101", "SOKMARKET")),
    ("Yemek", ("YEMEKSEPETI", "TRENDYOLYEMEK", "GETIRYEMEK", "POPEYES", "TAVUKDUNYASI", "LOKANTA", "RESTORAN", "CATERING", "GREENSALADS", "PEHLIVAN", "BUYUKEV", "MACGAL", "MERLAGIDA", "BURGERKING", "MCDONALDS", "KFC")),
    ("Kafe", ("KAFE", "CAFE", "COFFEE", "ESPRESSO", "MADO", "VANILPUFF", "STARBUCKS", "KAHVE", "MOC")),
    ("Abonelik / Dijital", ("APPLECOM", "YOUTUBE", "SPOTIFY", "NETFLIX", "GOOGLE", "DISNEY", "BLUTV", "EXXEN")),
    ("E-Ticaret", ("TRENDYOLCOM", "HEPSIBURADA", "AMAZON", "N11COM")),
    ("Giyim", ("ZARA", "LCWAIKIKI", "BERSHKA", "YILDIZSTORE", "PULLANDBEAR", "MAVI")),
    ("Kişisel Bakım", ("GRATIS", "WATSONS", "ROSSMANN", "SEPHORA")),
    ("Sağlık", ("ECZANE", "HASTANE", "MEDIKAL")),
    ("Tütün", ("TOBACCO", "TEKEL")),
]


@dataclass(frozen=True)
class AnalysisBundle:
    raw_transactions: pd.DataFrame
    analysis_data: pd.DataFrame
    transaction_type_summary: pd.DataFrame
    category_summary: pd.DataFrame
    monthly_summary: pd.DataFrame
    latest_month: str
    latest_month_data: pd.DataFrame
    latest_month_total: float
    latest_month_category_summary: pd.DataFrame
    spending_findings: pd.DataFrame
    financial_commentary: str
    report_path: str


def normalize_text(value: Any) -> str:
    if pd.isna(value):
        return ""
    text = unicodedata.normalize("NFKD", str(value))
    return "".join(c for c in text if not unicodedata.combining(c)).upper().strip()


def compact_text(value: Any) -> str:
    return re.sub(r"[\s/.\-_*]+", "", normalize_text(value))


def parse_turkish_amount(value: Any) -> float | None:
    if pd.isna(value):
        return None
    text = str(value).strip().replace("₺", "").replace("TL", "").replace(" ", "")
    text = text.replace(".", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def classify_transaction_type(description: str, amount: float) -> str:
    normalized, compact = normalize_text(description), compact_text(description)
    if "HESAPTANAKTARIM" in compact or "ODEME" in normalized:
        return "Ödeme"
    if "MAXIPUAN" in compact or "PUANILAVE" in compact:
        return "Puan İşlemi"
    if amount < 0:
        return "İade / İptal"
    if any(k in normalized for k in ("FAIZ", "MASRAF", "KOMISYON")):
        return "Faiz / Masraf"
    return "Harcama"


def categorize_transaction(description: str) -> str:
    normalized, compact = normalize_text(description), compact_text(description)
    for category, keywords in CATEGORY_RULES:
        if any(k in compact or k in normalized for k in keywords):
            return category
    return "Diğer"


def _extract_transaction_from_line(line: str, page_number: int) -> dict[str, Any] | None:
    date_match = re.match(r"^(\d{2}/\d{2}/\d{4})\s+(.+)$", line)
    if not date_match:
        return None
    date_text, remainder = date_match.group(1), date_match.group(2)
    amount_match = re.search(r"(?<![\d.,])-?(?:\d{1,3}(?:\.\d{3})+|\d+),\d{2}(?![\d.,])", remainder)
    if not amount_match:
        return None
    description = remainder[: amount_match.start()].strip()
    amount = parse_turkish_amount(amount_match.group())
    if not description or amount is None:
        return None
    return {
        "Sayfa": page_number,
        "Tarih": date_text,
        "Açıklama": description,
        "Tutar": amount,
        "Ek_Bilgi": remainder[amount_match.end():].strip(),
        "Islem_Tipi": classify_transaction_type(description, amount),
    }


def read_credit_card_statement(pdf_path: str | os.PathLike[str]) -> pd.DataFrame:
    pdf_path = str(pdf_path)
    if not os.path.exists(pdf_path):
        raise FileNotFoundError(f"Dosya bulunamadı: {pdf_path}")
    rows: list[dict[str, Any]] = []
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page_number, page in enumerate(pdf.pages, start=1):
                for raw_line in (page.extract_text(x_tolerance=1, y_tolerance=3) or "").splitlines():
                    transaction = _extract_transaction_from_line(" ".join(raw_line.split()), page_number)
                    if transaction:
                        rows.append(transaction)
    except Exception as exc:
        raise StatementParseError(f"PDF okunamadı: {exc}") from exc
    frame = pd.DataFrame(rows, columns=TRANSACTION_COLUMNS)
    if frame.empty:
        raise StatementParseError("PDF içerisinden işlem satırı çıkarılamadı. Dosya taranmış olabilir veya ekstre formatı henüz desteklenmiyor olabilir.")
    frame["Tarih"] = pd.to_datetime(frame["Tarih"], format="%d/%m/%Y", errors="coerce")
    frame = frame.dropna(subset=["Tarih"]).sort_values("Tarih").reset_index(drop=True)
    if frame.empty:
        raise StatementParseError("İşlem tarihleri okunamadı.")
    return frame


def build_analysis_tables(raw: pd.DataFrame) -> dict[str, Any]:
    data = raw[raw["Islem_Tipi"].isin(["Harcama", "Faiz / Masraf"])].copy()
    if data.empty:
        raise StatementParseError("Analize dahil edilebilecek harcama işlemi bulunamadı.")
    data["Tutar"] = data["Tutar"].abs()
    data["Kategori"] = data["Açıklama"].apply(categorize_transaction)
    data["Ay"] = data["Tarih"].dt.to_period("M").astype(str)

    type_summary = raw.groupby("Islem_Tipi", as_index=False).agg(Toplam_Tutar=("Tutar", "sum"), Islem_Sayisi=("Tutar", "count"), Ortalama_Tutar=("Tutar", "mean"))
    type_summary["Ortalama_Tutar"] = type_summary["Ortalama_Tutar"].round(2)
    category_summary = data.groupby("Kategori", as_index=False).agg(Toplam_Harcama=("Tutar", "sum"), Islem_Sayisi=("Tutar", "count"), Ortalama_Harcama=("Tutar", "mean")).sort_values("Toplam_Harcama", ascending=False).reset_index(drop=True)
    category_summary["Ortalama_Harcama"] = category_summary["Ortalama_Harcama"].round(2)
    monthly_summary = data.groupby("Ay", as_index=False).agg(Toplam_Harcama=("Tutar", "sum"), Islem_Sayisi=("Tutar", "count"), Ortalama_Islem_Tutari=("Tutar", "mean")).sort_values("Ay").reset_index(drop=True)
    monthly_summary["Ortalama_Islem_Tutari"] = monthly_summary["Ortalama_Islem_Tutari"].round(2)

    latest_month = str(data["Ay"].max())
    latest_data = data[data["Ay"] == latest_month].copy()
    latest_total = float(latest_data["Tutar"].sum())
    latest_categories = latest_data.groupby("Kategori", as_index=False).agg(Toplam_Harcama=("Tutar", "sum"), Islem_Sayisi=("Tutar", "count"), Ortalama_Harcama=("Tutar", "mean")).sort_values("Toplam_Harcama", ascending=False).reset_index(drop=True)
    latest_categories["Ortalama_Harcama"] = latest_categories["Ortalama_Harcama"].round(2)
    latest_categories["Harcama_Orani_%"] = (latest_categories["Toplam_Harcama"] / latest_total * 100).round(2)

    return {"analysis_data": data, "transaction_type_summary": type_summary, "category_summary": category_summary, "monthly_summary": monthly_summary, "latest_month": latest_month, "latest_month_data": latest_data, "latest_month_total": latest_total, "latest_month_category_summary": latest_categories}


def detect_spending_patterns(data: pd.DataFrame, summary: pd.DataFrame, total: float) -> pd.DataFrame:
    findings: list[dict[str, Any]] = []
    average = float(data["Tutar"].mean())
    def add(title: str, category: str, explanation: str, amount: float, risk: str) -> None:
        findings.append({"Tespit": title, "Kategori": category, "Açıklama": explanation, "Tutar": round(float(amount), 2), "Risk_Seviyesi": risk})
    def row_for(name: str) -> pd.DataFrame:
        return summary[summary["Kategori"] == name]

    top = summary.iloc[0]
    if top["Harcama_Orani_%"] >= 15:
        add("Yoğunlaşmış harcama kategorisi", str(top["Kategori"]), f"En yüksek kategori {top['Kategori']} ve toplam harcamanın %{top['Harcama_Orani_%']} kısmını oluşturuyor.", top["Toplam_Harcama"], "Orta")

    food = row_for("Yemek")
    if not food.empty:
        row = food.iloc[0]
        risk = "Yüksek" if row["Harcama_Orani_%"] >= 15 else "Orta" if row["Harcama_Orani_%"] >= 8 or row["Islem_Sayisi"] >= 5 else ""
        if risk:
            add("Yemek harcamaları dikkat çekiyor", "Yemek", f"{int(row['Islem_Sayisi'])} işlem, toplam harcamanın %{row['Harcama_Orani_%']} kısmı.", row["Toplam_Harcama"], risk)

    cafe = row_for("Kafe")
    if not cafe.empty:
        row = cafe.iloc[0]
        if row["Islem_Sayisi"] >= 4 or row["Harcama_Orani_%"] >= 6:
            add("Kafe harcamaları izlenmeli", "Kafe", f"{int(row['Islem_Sayisi'])} işlem ve %{row['Harcama_Orani_%']} pay.", row["Toplam_Harcama"], "Düşük / Orta")

    subscriptions = data[data["Kategori"] == "Abonelik / Dijital"]
    brands = subscriptions["Açıklama"].dropna().unique()
    if len(brands) >= 2:
        add("Birden fazla dijital abonelik", "Abonelik / Dijital", f"{len(brands)} farklı dijital ödeme görüldü: {', '.join(map(str, brands))}", subscriptions["Tutar"].sum(), "Orta")

    small = data[data["Tutar"] < 300]
    if len(small) >= 8:
        small_total = float(small["Tutar"].sum())
        add("Küçük ama sık harcamalar", "Genel", f"300 TL altı {len(small)} işlem toplam {small_total:.2f} TL ve toplamın %{small_total / total * 100:.2f} kısmı.", small_total, "Düşük / Orta")

    ecommerce = row_for("E-Ticaret")
    if not ecommerce.empty:
        row = ecommerce.iloc[0]
        if row["Harcama_Orani_%"] >= 10 or row["Islem_Sayisi"] >= 3:
            add("E-ticaret harcamaları izlenmeli", "E-Ticaret", f"{int(row['Islem_Sayisi'])} işlem ve %{row['Harcama_Orani_%']} pay.", row["Toplam_Harcama"], "Orta")

    interest = row_for("Faiz / Masraf")
    if not interest.empty and interest.iloc[0]["Toplam_Harcama"] > 0:
        row = interest.iloc[0]
        add("Faiz / masraf oluşmuş", "Faiz / Masraf", f"{int(row['Islem_Sayisi'])} işlem, toplam {row['Toplam_Harcama']:.2f} TL.", row["Toplam_Harcama"], "Yüksek")

    for _, row in summary[summary["Ortalama_Harcama"] > average * 1.25].iterrows():
        add("Ortalama üstü işlem tutarı", str(row["Kategori"]), f"Ortalama işlem {row['Ortalama_Harcama']:.2f} TL ve genel ortalamanın üzerinde.", row["Toplam_Harcama"], "Düşük / Orta")

    return pd.DataFrame(findings, columns=["Tespit", "Kategori", "Açıklama", "Tutar", "Risk_Seviyesi"])


def format_try(amount: float) -> str:
    return f"{amount:,.2f} TL".replace(",", "X").replace(".", ",").replace("X", ".")


def generate_financial_commentary(month: str, data: pd.DataFrame, total: float, summary: pd.DataFrame, findings: pd.DataFrame) -> str:
    top = summary.iloc[0]
    lines = ["AI Finans Asistanı - Aylık Finans Yorumu", "=" * 48, "", f"Analiz edilen ay: {month}", f"Toplam harcama: {format_try(total)}", f"Toplam işlem sayısı: {len(data)}", f"Ortalama işlem tutarı: {format_try(float(data['Tutar'].mean()))}", "", "Genel Değerlendirme", "-" * 20, f"En yüksek kategori {top['Kategori']}; {format_try(float(top['Toplam_Harcama']))} ve toplamın %{top['Harcama_Orani_%']} kısmı.", "", "Dikkat Edilmesi Gereken Alanlar", "-" * 30]
    if findings.empty:
        lines.append("Belirgin bir riskli harcama davranışı tespit edilmedi.")
    else:
        lines.append(f"{len(findings)} adet dikkat edilmesi gereken davranış tespit edildi.")
        for _, row in findings.iterrows():
            lines.extend(["", f"- {row['Tespit']}", f"  Kategori: {row['Kategori']}", f"  Açıklama: {row['Açıklama']}", f"  Risk: {row['Risk_Seviyesi']}"])
    lines.extend(["", "Not", "-" * 3, "Bu yorum şeffaf, kural tabanlı bir analiz motoru tarafından üretilmiştir ve finansal danışmanlık değildir."])
    return "\n".join(lines)


def _write_excel_report(path: str, raw: pd.DataFrame, tables: dict[str, Any], findings: pd.DataFrame, commentary: str) -> None:
    top10 = tables["latest_month_data"].sort_values("Tutar", ascending=False).head(10)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        raw.to_excel(writer, "PDF_Ham_Islemler", index=False)
        tables["analysis_data"].to_excel(writer, "Analiz_Verisi", index=False)
        tables["transaction_type_summary"].to_excel(writer, "Islem_Tipi_Ozeti", index=False)
        tables["category_summary"].to_excel(writer, "Kategori_Ozeti", index=False)
        tables["monthly_summary"].to_excel(writer, "Aylik_Ozet", index=False)
        tables["latest_month_category_summary"].to_excel(writer, "Son_Ay_Kategori", index=False)
        findings.to_excel(writer, "Harcama_Tespitleri", index=False)
        top10.to_excel(writer, "En_Yuksek_10_Islem", index=False)
        pd.DataFrame({"Finans_Yorumu": commentary.splitlines()}).to_excel(writer, "Finans_Yorumu", index=False)


def _add_excel_dashboard(path: str, tables: dict[str, Any], findings: pd.DataFrame) -> None:
    wb = load_workbook(path)
    if "Dashboard" in wb.sheetnames:
        del wb["Dashboard"]
    ws = wb.create_sheet("Dashboard", 0)
    ws["A1"] = "AI Finans Asistanı - PDF Ekstre Dashboard"
    ws["A1"].font = Font(size=18, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A1:H1")
    blue, light, green, orange, red = [PatternFill("solid", fgColor=c) for c in ("D9EAF7", "F2F2F2", "D9EAD3", "FCE5CD", "F4CCCC")]
    border = Border(**{side: Side(style="thin", color="D9D9D9") for side in ("left", "right", "top", "bottom")})
    latest, latest_data, monthly = tables["latest_month_category_summary"], tables["latest_month_data"], tables["monthly_summary"]
    overview = [("Analiz Edilen Ay", tables["latest_month"]), ("Toplam Harcama", tables["latest_month_total"]), ("Toplam İşlem Sayısı", len(latest_data)), ("Ortalama İşlem Tutarı", round(float(latest_data["Tutar"].mean()), 2)), ("Tespit Sayısı", len(findings)), ("En Yüksek Kategori", latest.iloc[0]["Kategori"])]
    for r, (label, value) in enumerate(overview, 3):
        ws[f"A{r}"], ws[f"B{r}"] = label, value
        ws[f"A{r}"].font, ws[f"A{r}"].fill, ws[f"B{r}"].fill = Font(bold=True), blue, light
        ws[f"A{r}"].border = ws[f"B{r}"].border = border
    ws["B4"].number_format = ws["B6"].number_format = '#,##0.00 "TL"'

    start = 11
    headers = ["Kategori", "Toplam Harcama", "İşlem Sayısı", "Ortalama Harcama", "Harcama Oranı %"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(start, c, h); cell.fill, cell.font, cell.border = green, Font(bold=True), border
    for i, row in latest.iterrows():
        r = start + 1 + i
        for c, value in enumerate([row["Kategori"], row["Toplam_Harcama"], row["Islem_Sayisi"], row["Ortalama_Harcama"], row["Harcama_Orani_%"]], 1):
            ws.cell(r, c, value).border = border
        ws[f"B{r}"].number_format = ws[f"D{r}"].number_format = '#,##0.00 "TL"'
    last = start + len(latest)
    bar = BarChart(); bar.title = "Kategori Bazlı Harcama"; bar.add_data(Reference(ws, min_col=2, min_row=start, max_row=last), titles_from_data=True); bar.set_categories(Reference(ws, min_col=1, min_row=start + 1, max_row=last)); bar.height, bar.width = 8, 16; ws.add_chart(bar, "G3")
    pie = PieChart(); pie.title = "Kategori Harcama Oranı"; pie.add_data(Reference(ws, min_col=5, min_row=start, max_row=last), titles_from_data=True); pie.set_categories(Reference(ws, min_col=1, min_row=start + 1, max_row=last)); pie.height, pie.width = 8, 12; ws.add_chart(pie, "G20")

    mstart = last + 4
    for c, h in enumerate(["Ay", "Toplam Harcama", "İşlem Sayısı", "Ortalama İşlem Tutarı"], 1):
        cell = ws.cell(mstart, c, h); cell.fill, cell.font, cell.border = orange, Font(bold=True), border
    for i, row in monthly.iterrows():
        r = mstart + 1 + i
        for c, value in enumerate([row["Ay"], row["Toplam_Harcama"], row["Islem_Sayisi"], row["Ortalama_Islem_Tutari"]], 1):
            ws.cell(r, c, value).border = border
    mlast = mstart + len(monthly)
    line = LineChart(); line.title = "Aylık Harcama Trendi"; line.add_data(Reference(ws, min_col=2, min_row=mstart, max_row=mlast), titles_from_data=True); line.set_categories(Reference(ws, min_col=1, min_row=mstart + 1, max_row=mlast)); line.height, line.width = 8, 16; ws.add_chart(line, "G37")

    fstart = mlast + 4
    for c, h in enumerate(["Tespit", "Kategori", "Açıklama", "Tutar", "Risk Seviyesi"], 1):
        cell = ws.cell(fstart, c, h); cell.fill, cell.font, cell.border = red, Font(bold=True), border
    for i, row in findings.reset_index(drop=True).iterrows():
        r = fstart + 1 + i
        for c, value in enumerate([row["Tespit"], row["Kategori"], row["Açıklama"], row["Tutar"], row["Risk_Seviyesi"]], 1):
            cell = ws.cell(r, c, value); cell.border = border; cell.alignment = Alignment(wrap_text=True, vertical="top")
    for col, width in {"A": 30, "B": 22, "C": 75, "D": 22, "E": 18, "F": 4, "G": 22, "H": 22, "I": 22, "J": 22, "K": 22, "L": 22}.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A11"
    wb.save(path)


def analyze_statement(pdf_path: str | os.PathLike[str], output_dir: str | os.PathLike[str] | None = None) -> AnalysisBundle:
    raw = read_credit_card_statement(pdf_path)
    tables = build_analysis_tables(raw)
    findings = detect_spending_patterns(tables["latest_month_data"], tables["latest_month_category_summary"], tables["latest_month_total"])
    commentary = generate_financial_commentary(tables["latest_month"], tables["latest_month_data"], tables["latest_month_total"], tables["latest_month_category_summary"], findings)
    output = Path(pdf_path).resolve().parent if output_dir is None else Path(output_dir).resolve()
    output.mkdir(parents=True, exist_ok=True)
    report = output / f"AI_Finans_Asistani_Raporu_{tables['latest_month']}.xlsx"
    _write_excel_report(str(report), raw, tables, findings, commentary)
    _add_excel_dashboard(str(report), tables, findings)
    return AnalysisBundle(raw, tables["analysis_data"], tables["transaction_type_summary"], tables["category_summary"], tables["monthly_summary"], tables["latest_month"], tables["latest_month_data"], tables["latest_month_total"], tables["latest_month_category_summary"], findings, commentary, str(report))
