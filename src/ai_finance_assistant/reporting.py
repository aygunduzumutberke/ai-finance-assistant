from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def write_excel_report(
    report_path: str | Path,
    raw_transactions: pd.DataFrame,
    tables: dict[str, Any],
    findings: pd.DataFrame,
    recurring: pd.DataFrame,
    anomalies: pd.DataFrame,
    budget_summary: pd.DataFrame,
    health_score: int,
    savings_potential: float,
    commentary: str,
) -> None:
    report_path = str(report_path)
    top_transactions = (
        tables["latest_month_data"].sort_values("Tutar", ascending=False).head(15)
    )
    overview = pd.DataFrame(
        {
            "Metrik": [
                "Analiz Edilen Ay",
                "Toplam Harcama",
                "İşlem Sayısı",
                "Finansal Sağlık Puanı",
                "Tahmini Tasarruf Potansiyeli",
                "Önceki Aya Göre Değişim %",
            ],
            "Değer": [
                tables["latest_month"],
                tables["latest_month_total"],
                len(tables["latest_month_data"]),
                health_score,
                savings_potential,
                tables["previous_month_change_pct"],
            ],
        }
    )

    with pd.ExcelWriter(report_path, engine="openpyxl") as writer:
        overview.to_excel(writer, "Genel_Ozet", index=False)
        raw_transactions.to_excel(writer, "PDF_Ham_Islemler", index=False)
        tables["analysis_data"].to_excel(writer, "Analiz_Verisi", index=False)
        tables["transaction_type_summary"].to_excel(writer, "Islem_Tipi_Ozeti", index=False)
        tables["category_summary"].to_excel(writer, "Kategori_Ozeti", index=False)
        tables["monthly_summary"].to_excel(writer, "Aylik_Ozet", index=False)
        tables["latest_month_category_summary"].to_excel(writer, "Son_Ay_Kategori", index=False)
        budget_summary.to_excel(writer, "Butce_Takibi", index=False)
        recurring.to_excel(writer, "Duzenli_Odemeler", index=False)
        anomalies.to_excel(writer, "Anomaliler", index=False)
        findings.to_excel(writer, "Harcama_Tespitleri", index=False)
        top_transactions.to_excel(writer, "En_Yuksek_15_Islem", index=False)
        pd.DataFrame({"Finans_Yorumu": commentary.splitlines()}).to_excel(
            writer, "Finans_Yorumu", index=False
        )

    add_excel_dashboard(
        report_path,
        tables=tables,
        findings=findings,
        budget_summary=budget_summary,
        health_score=health_score,
        savings_potential=savings_potential,
    )


def add_excel_dashboard(
    report_path: str | Path,
    tables: dict[str, Any],
    findings: pd.DataFrame,
    budget_summary: pd.DataFrame,
    health_score: int,
    savings_potential: float,
) -> None:
    workbook = load_workbook(report_path)
    if "Dashboard" in workbook.sheetnames:
        del workbook["Dashboard"]
    sheet = workbook.create_sheet("Dashboard", 0)

    sheet["A1"] = "AI Finans Asistanı V2 - Dashboard"
    sheet["A1"].font = Font(size=19, bold=True)
    sheet["A1"].alignment = Alignment(horizontal="center")
    sheet.merge_cells("A1:H1")

    blue = PatternFill("solid", fgColor="D9EAF7")
    gray = PatternFill("solid", fgColor="F2F2F2")
    green = PatternFill("solid", fgColor="D9EAD3")
    orange = PatternFill("solid", fgColor="FCE5CD")
    red = PatternFill("solid", fgColor="F4CCCC")
    border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    latest = tables["latest_month_category_summary"]
    latest_data = tables["latest_month_data"]
    previous_change = tables["previous_month_change_pct"]
    metrics = [
        ("Analiz Edilen Ay", tables["latest_month"]),
        ("Toplam Harcama", tables["latest_month_total"]),
        ("Toplam İşlem Sayısı", len(latest_data)),
        ("Ortalama İşlem", float(latest_data["Tutar"].mean())),
        ("Finansal Sağlık Puanı", health_score),
        ("Tasarruf Potansiyeli", savings_potential),
        ("Önceki Aya Göre Değişim %", previous_change if previous_change is not None else "Veri yok"),
        ("Tespit Sayısı", len(findings)),
    ]
    for row_index, (label, value) in enumerate(metrics, start=3):
        sheet[f"A{row_index}"] = label
        sheet[f"B{row_index}"] = value
        sheet[f"A{row_index}"].font = Font(bold=True)
        sheet[f"A{row_index}"].fill = blue
        sheet[f"B{row_index}"].fill = gray
        sheet[f"A{row_index}"].border = border
        sheet[f"B{row_index}"].border = border
    for cell in ("B4", "B6", "B8"):
        sheet[cell].number_format = '#,##0.00 "TL"'

    category_start = 13
    headers = ["Kategori", "Toplam Harcama", "İşlem Sayısı", "Ortalama", "Oran %"]
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(category_start, column, header)
        cell.fill, cell.font, cell.border = green, Font(bold=True), border

    for offset, row in latest.reset_index(drop=True).iterrows():
        excel_row = category_start + 1 + offset
        values = [
            row["Kategori"], row["Toplam_Harcama"], row["Islem_Sayisi"],
            row["Ortalama_Harcama"], row["Harcama_Orani_%"],
        ]
        for column, value in enumerate(values, start=1):
            sheet.cell(excel_row, column, value).border = border
        sheet[f"B{excel_row}"].number_format = '#,##0.00 "TL"'
        sheet[f"D{excel_row}"].number_format = '#,##0.00 "TL"'

    category_last = category_start + len(latest)
    bar = BarChart()
    bar.title = "Kategori Bazlı Harcama"
    bar.add_data(
        Reference(sheet, min_col=2, min_row=category_start, max_row=category_last),
        titles_from_data=True,
    )
    bar.set_categories(
        Reference(sheet, min_col=1, min_row=category_start + 1, max_row=category_last)
    )
    bar.height, bar.width = 8, 16
    sheet.add_chart(bar, "G3")

    pie = PieChart()
    pie.title = "Kategori Harcama Oranı"
    pie.add_data(
        Reference(sheet, min_col=5, min_row=category_start, max_row=category_last),
        titles_from_data=True,
    )
    pie.set_categories(
        Reference(sheet, min_col=1, min_row=category_start + 1, max_row=category_last)
    )
    pie.height, pie.width = 8, 12
    sheet.add_chart(pie, "G20")

    monthly = tables["monthly_summary"]
    monthly_start = category_last + 4
    for column, header in enumerate(
        ["Ay", "Toplam Harcama", "İşlem Sayısı", "Ortalama İşlem"], start=1
    ):
        cell = sheet.cell(monthly_start, column, header)
        cell.fill, cell.font, cell.border = orange, Font(bold=True), border
    for offset, row in monthly.reset_index(drop=True).iterrows():
        excel_row = monthly_start + 1 + offset
        values = [row["Ay"], row["Toplam_Harcama"], row["Islem_Sayisi"], row["Ortalama_Islem_Tutari"]]
        for column, value in enumerate(values, start=1):
            sheet.cell(excel_row, column, value).border = border
    monthly_last = monthly_start + len(monthly)

    line = LineChart()
    line.title = "Aylık Harcama Trendi"
    line.add_data(
        Reference(sheet, min_col=2, min_row=monthly_start, max_row=monthly_last),
        titles_from_data=True,
    )
    line.set_categories(
        Reference(sheet, min_col=1, min_row=monthly_start + 1, max_row=monthly_last)
    )
    line.height, line.width = 8, 16
    sheet.add_chart(line, "G37")

    budget_start = monthly_last + 4
    for column, header in enumerate(
        ["Kategori", "Bütçe", "Harcama", "Kalan", "Kullanım %", "Durum"], start=1
    ):
        cell = sheet.cell(budget_start, column, header)
        cell.fill, cell.font, cell.border = red, Font(bold=True), border
    for offset, row in budget_summary.reset_index(drop=True).iterrows():
        excel_row = budget_start + 1 + offset
        values = [row["Kategori"], row["Butce"], row["Harcama"], row["Kalan"], row["Kullanim_%"], row["Durum"]]
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(excel_row, column, value)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True)

    for column, width in {
        "A": 31, "B": 22, "C": 25, "D": 21, "E": 18, "F": 18,
        "G": 22, "H": 22, "I": 22, "J": 22, "K": 22, "L": 22,
    }.items():
        sheet.column_dimensions[column].width = width
    sheet.freeze_panes = "A13"
    workbook.save(report_path)
